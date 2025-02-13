import xml.etree.ElementTree as ET
import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from tkinter import Tk, Label, Button, filedialog, StringVar, Entry, Toplevel
from tkinter import ttk
import webbrowser
from datetime import datetime

def process_xml_to_xlsx(xml_file, output_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Processed Data"

    headers = [
        "Descrição (ID + Nome)", "Quantidade", "COMP", "LARG", 
        "BORDA SUP", "BORDA INF", "BORDA DIR", "BORDA ESQ", "Material", "Observações"
    ]
    sheet.append(headers)

    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    for col in sheet[1]:
        col.fill = header_fill
        col.font = header_font
        col.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = [30, 15, 10, 10, 10, 10, 10, 10, 25, 40]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width

    for item in root.findall(".//ITEM"):
        unique_id = item.get("UNIQUEPARENTID", "")
        description = item.get("DESCRIPTION", "")

        try:
            if unique_id.startswith("-") or abs(int(unique_id)) < 10:
                unique_id = item.get("UNIQUEID", unique_id)
        except ValueError:
            pass

        combined_description = f"{unique_id} - {description}" if unique_id and description else ""

        quantity = item.get("QUANTITY", "0")
        rounded_quantity = math.ceil(float(quantity))

        text_dimension = item.get("TEXTDIMENSION", "")
        dimensions = text_dimension.replace(",", ".").split("x")

        comp = ""
        larg = ""
        if len(dimensions) > 0:
            try:
                comp_value = float(dimensions[0].strip())
                comp = int(comp_value) if comp_value % 1 < 0.5 else int(comp_value) + 1
            except ValueError:
                comp = ""
        if len(dimensions) > 1:
            try:
                larg_value = float(dimensions[-1].strip())
                larg = int(larg_value) if larg_value % 1 < 0.5 else int(larg_value) + 1
            except ValueError:
                larg = ""

        borda_sup = "x" if item.find(".//FITA_BORDA_1[@REFERENCE='1']") is not None else ""
        borda_inf = "x" if item.find(".//FITA_BORDA_2[@REFERENCE='1']") is not None else ""
        borda_dir = "x" if item.find(".//FITA_BORDA_3[@REFERENCE='1']") is not None else ""
        borda_esq = "x" if item.find(".//FITA_BORDA_4[@REFERENCE='1']") is not None else ""

        reference = item.get("REFERENCE", "")
        material_info = reference.split(".")
        material_full = " ".join(reversed(material_info)) if material_info else ""

        if "MDF" not in material_full and "porta" not in description.lower():
            continue

        material_parts = material_full.split()
        material = " ".join(material_parts[:3]) if len(material_parts) >= 3 else material_full

        observations = item.get("OBSERVATIONS", "")

        sheet.append([
            combined_description, rounded_quantity, comp, larg, 
            borda_sup, borda_inf, borda_dir, borda_esq, material, observations
        ])

    workbook.save(output_file)

def browse_xml():
    file_path = filedialog.askopenfilename(filetypes=[("XML files", "*.xml")])
    if file_path:
        xml_path.set(file_path)

def save_xlsx():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        xlsx_path.set(file_path)

def clear_fields():
    xml_path.set("")
    xlsx_path.set("")
    result_label.config(text="")

def convert_file():
    xml_file = xml_path.get()
    xlsx_file = xlsx_path.get()

    if not xml_file or not xlsx_file:
        result_label.config(text="Por favor, selecione os arquivos de entrada e saída.")
        return

    try:
        process_xml_to_xlsx(xml_file, xlsx_file)
        result_label.config(text="Conversão concluída com sucesso!")
    except Exception as e:
        result_label.config(text=f"Erro: {e}")

def open_website():
    webbrowser.open("https://enzomartinsdev.com")

def show_conversion_screen():
    root = Tk()
    root.title("Conversor XML para XLSX")
    root.geometry("600x400")
    root.configure(bg="#f0f0f0")

    global xml_path, xlsx_path, result_label
    xml_path = StringVar()
    xlsx_path = StringVar()

    Label(root, text="Bem-vindo ao Conversor XML para XLSX", font=("Arial", 16, "bold"), bg="#f0f0f0").pack(pady=10)
    Button(root, text="Selecionar XML", command=browse_xml, bg="#0078D7", fg="white", font=("Arial", 12), padx=10).pack(pady=10)
    Entry(root, textvariable=xml_path, width=50, font=("Arial", 10)).pack(pady=5)
    Button(root, text="Salvar como XLSX", command=save_xlsx, bg="#0078D7", fg="white", font=("Arial", 12), padx=10).pack(pady=10)
    Entry(root, textvariable=xlsx_path, width=50, font=("Arial", 10)).pack(pady=5)
    Button(root, text="Converter", command=convert_file, bg="#28A745", fg="white", font=("Arial", 12), padx=10).pack(pady=10)
    Button(root, text="Fazer Nova Conversão", command=clear_fields, bg="#FFC107", fg="black", font=("Arial", 12), padx=10).pack(pady=10)

    result_label = Label(root, text="", fg="#333333", bg="#f0f0f0", font=("Arial", 12))
    result_label.pack(pady=5)

    link_label = Label(root, text="Desenvolvido por Enzo Martins", fg="#0078D7", cursor="hand2", bg="#f0f0f0", font=("Arial", 12))
    link_label.pack(pady=30)
    link_label.bind("<Button-1>", lambda e: open_website())

    root.mainloop()

def login_screen():
    def attempt_login():
        username = username_var.get()
        password = password_var.get()

        if username == "Wesley" and password == "wesley123":
            login_window.destroy()
            show_conversion_screen()
        else:
            login_result.config(text="Usuário ou senha inválidos, ou login expirado!", fg="red")

    login_window = Tk()
    login_window.title("Login")
    login_window.geometry("300x200")

    Label(login_window, text="Usuário:", font=("Arial", 12)).pack(pady=5)
    username_var = StringVar()
    Entry(login_window, textvariable=username_var, font=("Arial", 12)).pack(pady=5)

    Label(login_window, text="Senha:", font=("Arial", 12)).pack(pady=5)
    password_var = StringVar()
    Entry(login_window, textvariable=password_var, font=("Arial", 12), show="*").pack(pady=5)

    Button(login_window, text="Login", command=attempt_login, bg="#0078D7", fg="white", font=("Arial", 12)).pack(pady=10)

    login_result = Label(login_window, text="", font=("Arial", 10))
    login_result.pack(pady=5)

    login_window.mainloop()

login_screen()